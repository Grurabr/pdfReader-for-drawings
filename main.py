import shutil
import sys
import os
import math

import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import win32com.client as win32
import fitz
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QFileDialog, QLabel, QVBoxLayout, QPushButton, QWidget,
    QGraphicsView, QGraphicsScene, QGraphicsRectItem, QGraphicsEllipseItem, QGraphicsLineItem,
    QHBoxLayout, QTableWidget, QTableWidgetItem, QHeaderView, QDialog, QSizePolicy, QProgressBar, QLineEdit
)
from PyQt5.QtGui import QPixmap, QImage, QFont
from PyQt5.QtCore import Qt, QRectF, QEvent

class Instructions(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Ohjeet")
        self.setModal(False)
        self.setGeometry(1000, 200, 300, 400)

        self.layout = QVBoxLayout()
        self.setLayout(self.layout)

        self.label = QLabel("1 Valitse PDF-tiedosto käsittelyä varten\n\n"
                            "2 Valitse mitat, jotka haluat lisätä jokaiselle tiedoston sivulle\n\n"
                            "3 Siirrä ympyröitä PDF:ssä niin, etteivät ne peitä muuta tekstiä\n\n"
                            "4 Tee muutoksia taulukon tekstiin (muokkaa tekstiä tai \npoista tarpeeton rivi painamalla 'Delete selected block')\n\n"
                            "5 Luo uusi PDF ja Excel painamalla 'Print documents' -painiketta")
        self.layout.addWidget(self.label)

        self.ok_button = QPushButton("Ok")
        self.ok_button.clicked.connect(self.accept)
        self.layout.addWidget(self.ok_button)

class PrintDialog(QDialog):
    def __init__(self, parent=None, default_path=""):
        super().__init__()
        self.setWindowTitle("Print documents")
        self.setModal(True)  # Locks the main window
        self.setGeometry(400, 200, 500, 150)

        self.excel_malli_path = default_path

        # Layuot
        self.layout = QVBoxLayout()
        self.setLayout(self.layout)

        # Label
        self.label = QLabel("Excel malli:")
        self.layout.addWidget(self.label)

        self.path_field = QLineEdit()
        self.path_field.setPlaceholderText(f"{self.excel_malli_path}")
        if self.excel_malli_path:  # Если задан путь по умолчанию, отображаем его
            self.path_field.setText(self.excel_malli_path)
        self.layout.addWidget(self.path_field)

        # Кнопка выбора пути
        self.browse_button = QPushButton("Valitse malli")
        self.browse_button.clicked.connect(self.browse_file)
        self.layout.addWidget(self.browse_button)

        # Кнопка подтверждения
        self.ok_button = QPushButton("Ok")
        self.ok_button.clicked.connect(self.accept)
        self.layout.addWidget(self.ok_button)

        # Кнопка отмены
        self.cancel_button = QPushButton("Peruuta")
        self.cancel_button.clicked.connect(self.reject)
        self.layout.addWidget(self.cancel_button)

    def browse_file(self):
        # Открытие диалога выбора файла
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Valotse Excel malli",
            "",
            "Excel Files (*.xls *.xlsx)"
        )
        if file_path:
            self.path_field.setText(file_path)
            self.excel_malli_path = file_path

    def get_excel_path(self):
        return self.excel_malli_path





class ProgressWindow(QDialog):
    def __init__(self, parent=None):
        super().__init__()
        self.setWindowTitle("Processing")
        self.setModal(True) #Locks the main window
        self.setGeometry(400, 200, 300, 150)

        #Layuot
        self.layout = QVBoxLayout()
        self.setLayout(self.layout)

        # Label
        self.label = QLabel("Processing files...")
        self.layout.addWidget(self.label)

        # Progress bar
        self.progress_bar = QProgressBar(self)
        self.progress_bar.setMinimum(0)
        self.progress_bar.setMaximum(100)
        self.layout.addWidget(self.progress_bar)

        # Ok buttom
        self.ok_button = QPushButton("OK")
        self.ok_button.setEnabled(False)
        self.ok_button.clicked.connect(self.accept)
        self.layout.addWidget(self.ok_button)

    def update_progress(self, value, messege=""):
        self.progress_bar.setValue(value)
        if messege:
            self.label.setText(messege)

    def set_compledet(self):
        self.label.setText("Files saved successfully!")
        self.ok_button.setEnabled(True)

class PDFViewer(QMainWindow):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("PDF Viewer and Processor")
        self.setGeometry(100, 100, 1600, 900)

        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)

        # Main layout
        self.main_layout = QHBoxLayout(self.central_widget)

        # Left layout (PDF Viewer and Navigation)
        self.left_layout = QVBoxLayout()
        self.main_layout.addLayout(self.left_layout)

        # File open button
        self.open_button = QPushButton("Open PDF File")
        self.open_button.clicked.connect(self.open_pdf)
        self.left_layout.addWidget(self.open_button)

        # Graphics view for PDF rendering
        self.graphics_view = QGraphicsView()
        self.graphics_scene = QGraphicsScene()
        self.graphics_view.setScene(self.graphics_scene)
        self.left_layout.addWidget(self.graphics_view)

        # Instructions label
        self.instructions_label = QLabel("Click and drag to select an area.")
        self.instructions_label.setAlignment(Qt.AlignCenter)
        self.left_layout.addWidget(self.instructions_label)

        # Navigation buttons for pages
        nav_layout = QHBoxLayout()
        self.prev_button = QPushButton("Previous Page")
        self.prev_button.clicked.connect(self.prev_page)
        self.prev_button.setEnabled(False)  # Disabled initially
        nav_layout.addWidget(self.prev_button)

        self.next_button = QPushButton("Next Page")
        self.next_button.clicked.connect(self.next_page)
        self.next_button.setEnabled(False)  # Disabled initially
        nav_layout.addWidget(self.next_button)

        self.left_layout.addLayout(nav_layout)

        # Right layout (Text Blocks Table and Edit/Delete)
        self.right_layout = QVBoxLayout()
        self.main_layout.addLayout(self.right_layout)

        # Special characters buttons
        self.character_buttons_layout = QHBoxLayout()

        self.special_characters = ["Ø", "±", "°", "↧", "⊥", "||", "≈", "Ra"]
        total_width = 300
        button_width = total_width//len(self.special_characters) - 5

        for char in self.special_characters:
            button = QPushButton(char)
            button.setMaximumWidth(button_width)
            button.setMinimumWidth(button_width)
            button.setStyleSheet("font-size: 12px;")
            button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
            button.clicked.connect(lambda _, c=char: self.insert_character(c))
            self.character_buttons_layout.addWidget(button)



        spacer = QWidget()
        spacer.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.right_layout.addLayout(self.character_buttons_layout)

        # Table
        self.text_blocks_table = QTableWidget()
        self.text_blocks_table.setColumnCount(3)
        self.text_blocks_table.setHorizontalHeaderLabels(["Page", "Index", "Text"])
        self.text_blocks_table.verticalHeader().hide()
        self.text_blocks_table.setMaximumWidth(300)
        self.right_layout.addWidget(self.text_blocks_table)

        self.text_blocks_table.setColumnWidth(0, 30)
        self.text_blocks_table.setColumnWidth(1, 30)

        header = self.text_blocks_table.horizontalHeader()
        header.setSectionResizeMode(2, QHeaderView.Stretch)

        # Table text update
        self.text_blocks_table.cellChanged.connect(self.update_block_text)
        self.text_blocks_table.cellClicked.connect(self.highlight_selected_block)

        # self.text_blocks_table.setSortingEnabled(True)

        # Buttons
        self.add_row_button = QPushButton("Lisää rivi")
        self.add_row_button.clicked.connect(self.add_row_to_table)
        self.add_row_button.setMaximumWidth(300)
        self.right_layout.addWidget(self.add_row_button)

        self.instruction_button = QPushButton("Ohjeet")
        self.instruction_button.clicked.connect(self.instruction_block)
        self.instruction_button.setMaximumWidth(300)
        self.right_layout.addWidget(self.instruction_button)

        self.delete_button = QPushButton("Delete selected block")
        self.delete_button.clicked.connect(self.delete_block)
        self.delete_button.setMaximumWidth(300)
        self.right_layout.addWidget(self.delete_button)

        self.clear_button = QPushButton("Clear all")
        self.clear_button.clicked.connect(self.clear_all)
        self.clear_button.setMaximumWidth(300)
        self.right_layout.addWidget(self.clear_button)

        self.print_documents_button = QPushButton("Print documents")
        self.print_documents_button.clicked.connect(self.print_documents_block)
        self.print_documents_button.setMaximumWidth(300)
        self.print_documents_button.setEnabled(False)
        self.right_layout.addWidget(self.print_documents_button)

        self.blocks_data = []  # To store all extracted blocks with page and area info
        self.highlights = []  # To store QGraphicsRectItem objects for highlights
        self.pdf_document = None
        self.pdf_document_name = ""
        self.current_page = 0
        self.rect_item = None
        self.selection_start = None
        self.selection_rect = None
        self.dragged_circle = None  # Track the dragged circle
        self.measurement_number = [0]
        self.measurement_text = [0]
        self.radius = 12
        self.save_path = ""

        # Install event filter for graphics view
        self.graphics_view.viewport().installEventFilter(self)




    def insert_character(self, char):
        current_row = self.text_blocks_table.currentRow()
        current_column = self.text_blocks_table.currentColumn()

        if current_row is not None and current_column == 2:
            item = self.text_blocks_table.item(current_row, current_column)
            if item is None:
                # if cell empty, create new item
                item = QTableWidgetItem("")
                self.text_blocks_table.setItem(current_row, current_column, item)

            current_text = item.text()

            new_text = char + current_text

            item.setText(new_text)

    def highlight_selected_block(self, row):

        for _, circle_item, _, _ in self.highlights[self.current_page]:
            if circle_item.scene():
                pen = circle_item.pen()
                pen.setColor(Qt.red)
                pen.setWidth(1)
                circle_item.setPen(pen)

        if not self.text_blocks_table.item(row, 1).text().endswith("K"):
            block_index = int(self.text_blocks_table.item(row, 1).text())

            for rect_item, circle_item, line_item, text_item in self.highlights[self.current_page]:
                if circle_item.data(0) == block_index:
                    pen = circle_item.pen()
                    pen.setColor(Qt.green)
                    pen.setWidth(4)
                    circle_item.setPen(pen)
                    break



    def open_pdf(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Open PDF File", "", "PDF Files (*.pdf)")
        if file_path:

            self.pdf_document_name = os.path.splitext(os.path.basename(file_path))[0]

            self.blocks_data.clear()

            if self.highlights != []:
                for i in range(len(self.highlights)):
                    for rect_item, circle_item, line_item, text_item in self.highlights[i]:
                        self.graphics_scene.removeItem(rect_item)
                        self.graphics_scene.removeItem(circle_item)
                        self.graphics_scene.removeItem(line_item)
                        self.graphics_scene.removeItem(text_item)

                        self.highlights[i].clear()


            self.highlights.clear()

            self.update_blocks_table()

            self.load_pdf(file_path)

    def load_pdf(self, file_path):
        self.pdf_document = fitz.open(file_path)
        self.print_documents_button.setEnabled(True)

        # Enable navigation buttons if there are multiple pages
        if len(self.pdf_document) > 1:
            self.next_button.setEnabled(True)
        else:
            self.next_button.setEnabled(False)

        number_of_pages = self.pdf_document.page_count
        for i in range(number_of_pages):
            self.blocks_data.append([])
            self.highlights.append([])

        self.current_page = 0
        self.measurement_number = [0]
        self.render_page()

        self.prev_button.setEnabled(False)

    def render_page(self):
        if not self.pdf_document:
            return

        page = self.pdf_document.load_page(self.current_page)
        pix = page.get_pixmap()

        # Convert pixmap to QImage
        image = QImage(pix.samples, pix.width, pix.height, pix.stride, QImage.Format_RGB888)
        pixmap = QPixmap.fromImage(image)

        self.graphics_scene.clear()
        self.graphics_scene.addPixmap(pixmap)

        # Redraw highlights for the current page
        self.draw_highlights()

        # Set scene rect to match the image size
        self.graphics_scene.setSceneRect(QRectF(0, 0, pix.width, pix.height))

        self.instructions_label.setText(f"Page {self.current_page + 1}/{len(self.pdf_document)}. Click and drag to select an area.")

    def draw_highlights(self):
        for rect_item, circle_item, line_item, text_item in list(self.highlights[self.current_page]):
            try:
                if rect_item.scene():
                    self.graphics_scene.removeItem(rect_item)
                if circle_item.scene():
                    self.graphics_scene.removeItem(circle_item)
                if line_item.scene():
                    self.graphics_scene.removeItem(line_item)
                if text_item.scene():
                    self.graphics_scene.removeItem(text_item)
            except RuntimeError:
                pass

        self.highlights[self.current_page].clear()

        for block in self.blocks_data[self.current_page]:
            if block["index"] != None:
                if block["page"] == self.current_page:
                    x0, y0, x1, y1 = block["rect"]

                    # Draw rectangle around the block
                    rect_item = QGraphicsRectItem(x0, y0, x1 - x0, y1 - y0)
                    rect_item.setPen(Qt.blue)
                    self.graphics_scene.addItem(rect_item)

                    # Use saved circle position if available
                    if block["circle_position"]:
                        circle_x, circle_y = block["circle_position"]
                    else:
                        circle_x, circle_y = x1 + 20, y0 - 10

                    # Draw circle with index
                    circle_radius = self.radius
                    circle_item = QGraphicsEllipseItem(
                        circle_x - circle_radius, circle_y - circle_radius,
                        circle_radius * 2, circle_radius * 2
                    )
                    circle_item.setPen(Qt.red)
                    self.graphics_scene.addItem(circle_item)

                    # Draw line connecting rectangle and circle
                    new_lx, new_ly = self.shorten_line_to_circle(circle_x, circle_y, x1, y0, circle_radius)

                    line_item = QGraphicsLineItem(x1, y0, new_lx, new_ly)
                    line_item.setPen(Qt.blue)
                    self.graphics_scene.addItem(line_item)

                    # Add text inside the circle
                    text_item = self.graphics_scene.addText(str(block["index"]))
                    text_item.setDefaultTextColor(Qt.red)
                    text_item.setFont(QFont("Arial", 10))
                    text_item.setPos(circle_x - 7, circle_y - 12)

                    # Enable moving the circle
                    circle_item.setFlag(QGraphicsEllipseItem.ItemIsMovable, True)
                    circle_item.setFlag(QGraphicsEllipseItem.ItemSendsGeometryChanges, True)
                    circle_item.setData(0, block["index"])  # Store the index for reference

                    self.highlights[self.current_page].append((rect_item, circle_item, line_item, text_item))


    def prev_page(self):
        if self.current_page > 0:
            self.current_page -= 1
            self.render_page()

        # Update button states
        self.next_button.setEnabled(self.current_page < len(self.pdf_document) - 1)
        self.prev_button.setEnabled(self.current_page > 0)

    def next_page(self):
        if self.current_page < len(self.pdf_document) - 1:
            self.current_page += 1
            self.render_page()

        # Update button states
        self.next_button.setEnabled(self.current_page < len(self.pdf_document) - 1)
        self.prev_button.setEnabled(self.current_page > 0)

    def eventFilter(self, source, event):
        if source == self.graphics_view.viewport():
            if event.type() == QEvent.MouseButtonPress and event.button() == Qt.LeftButton:
                return self.handle_mouse_press(event)
            elif event.type() == QEvent.MouseMove and event.buttons() == Qt.LeftButton:
                return self.handle_mouse_move(event)
            elif event.type() == QEvent.MouseButtonRelease and event.button() == Qt.LeftButton:
                return self.handle_mouse_release(event)

        if event.type() == QEvent.FocusOut and source is self.text_blocks_table:

            self.suggestions_menu.hide()

        return super().eventFilter(source, event)

    def handle_mouse_press(self, event):
        view_pos = self.graphics_view.mapToScene(event.pos())


        # Check if a circle was clicked
        for _, circle_item, line_item, text_item in self.highlights[self.current_page]:
            #print(self.highlights[self.current_page])
            #print(circle_item)
            #print(view_pos)
            if circle_item.contains(view_pos):
                self.dragged_circle = (circle_item, line_item, text_item)
                return True

        # If not, start a new rectangle selection
        self.selection_start = view_pos

        try:
            if self.rect_item is not None:
                self.graphics_scene.removeItem(self.rect_item)
                self.rect_item = None
        except RuntimeError:
            self.rect_item = None

        return True

    def handle_mouse_move(self, event):
        if self.dragged_circle:
            # Move the circle and update line and text
            circle_item, line_item, text_item = self.dragged_circle
            scene_pos = self.graphics_view.mapToScene(event.pos())

            # Update circle position
            circle_item.setRect(
                scene_pos.x() - 12, scene_pos.y() - 12, 24, 24
            )

            # Update line position
            new_lx, new_ly = self.shorten_line_to_circle(scene_pos.x(), scene_pos.y(), line_item.line().x1(), line_item.line().y1(), 12)

            line_item.setLine(
                line_item.line().x1(), line_item.line().y1(),
                new_lx, new_ly
            )

            # Update text position
            text_item.setPos(scene_pos.x() - 7, scene_pos.y() - 12)

            # Save the new circle position to blocks_data
            index = circle_item.data(0)  # Get the block index from circle's data
            for block in self.blocks_data[self.current_page]:
                if block["index"] == index:
                    block["circle_position"] = (scene_pos.x(), scene_pos.y())
                    break

            return True

        if self.selection_start:
            view_pos = self.graphics_view.mapToScene(event.pos())
            x0, y0 = self.selection_start.x(), self.selection_start.y()
            x1, y1 = view_pos.x(), view_pos.y()

            rect = QRectF(min(x0, x1), min(y0, y1), abs(x1 - x0), abs(y1 - y0))
            if not self.rect_item:
                self.rect_item = QGraphicsRectItem(rect)
                self.rect_item.setPen(Qt.red)
                self.graphics_scene.addItem(self.rect_item)
            else:
                self.rect_item.setRect(rect)
        return True

    def handle_mouse_release(self, event):
        if self.dragged_circle:
            # Stop dragging the circle
            self.dragged_circle = None
            return True

        if self.selection_start:
            view_pos = self.graphics_view.mapToScene(event.pos())
            x0, y0 = self.selection_start.x(), self.selection_start.y()
            x1, y1 = view_pos.x(), view_pos.y()

            self.selection_rect = QRectF(min(x0, x1), min(y0, y1), abs(x1 - x0), abs(y1 - y0))
            # print(f"Selected area: {self.selection_rect}")

            # Extract and display text blocks from the selected area
            self.extract_text_blocks()

            self.selection_start = None
        return True

    def extract_text_blocks(self):
        if not self.pdf_document or not self.selection_rect:
            return

        # Convert selection rect to PDF coordinates
        page = self.pdf_document.load_page(self.current_page)
        #print(page)
        pdf_width, pdf_height = page.rect.width, page.rect.height
        x_scale = pdf_width / self.graphics_scene.sceneRect().width()
        y_scale = pdf_height / self.graphics_scene.sceneRect().height()

        x0 = self.selection_rect.left() * x_scale
        y0 = self.selection_rect.top() * y_scale
        x1 = self.selection_rect.right() * x_scale
        y1 = self.selection_rect.bottom() * y_scale

        selected_blocks = page.get_text("blocks")

        for block in selected_blocks:
            bx0, by0, bx1, by1, text = block[:5]
            if x0 <= bx0 <= x1 and y0 <= by0 <= y1:
                if not any(b["rect"] == (bx0, by0, bx1, by1) and b["page"] == self.current_page for b in self.blocks_data[self.current_page]):

                    self.blocks_data[self.current_page].append({
                        "page": self.current_page,
                        "rect": (bx0, by0, bx1, by1),
                        "text": text.replace("\n", " ").strip(),
                        "index": self.measurement_number_calc(),
                        "circle_position": None
                    })

        self.update_blocks_table()
        self.draw_highlights()

    def measurement_number_calc(self):
        self.measurement_number.append(self.measurement_number[-1]+1)
        return self.measurement_number[-1]

    def shorten_line_to_circle(self, cx, cy, lx, ly, radius):
        dx = lx - cx
        dy = ly - cy
        distance = (dx ** 2 + dy ** 2) ** 0.5

        if distance <= radius:
            return lx, ly

        k = radius / distance

        nx = cx + k * dx
        ny = cy + k * dy

        return nx, ny

    def update_blocks_table(self):
        self.text_blocks_table.setRowCount(0)
        for page_number, page_blocks in enumerate(self.blocks_data):
            for block in page_blocks:
                if block["index"] is not None:
                    row_position = self.text_blocks_table.rowCount()
                    self.text_blocks_table.insertRow(row_position)

                    page_item = QTableWidgetItem(str(page_number + 1))
                    page_item.setFont(QFont("Arial", weight=QFont.Bold))
                    page_item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                    self.text_blocks_table.setItem(row_position, 0, page_item)

                    index_item = QTableWidgetItem(str(block['index']))
                    index_item.setFont(QFont("Arial", weight=QFont.Bold))
                    index_item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                    self.text_blocks_table.setItem(row_position, 1, index_item)

                    text_item = QTableWidgetItem(block['text'])
                    self.text_blocks_table.setItem(row_position, 2, text_item)
                else:
                    row_position = self.text_blocks_table.rowCount()
                    self.text_blocks_table.insertRow(row_position)

                    page_item = QTableWidgetItem(str(page_number + 1))
                    page_item.setFont(QFont("Arial", weight=QFont.Bold))
                    page_item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                    self.text_blocks_table.setItem(row_position, 0, page_item)

                    index_item = QTableWidgetItem(str(block['text_index']))
                    index_item.setFont(QFont("Arial", weight=QFont.Bold))
                    #index_item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                    self.text_blocks_table.setItem(row_position, 1, index_item)

                    text_item = QTableWidgetItem(block['text'])
                    self.text_blocks_table.setItem(row_position, 2, text_item)



    def delete_block(self):
        current_row = self.text_blocks_table.currentRow()
        if current_row >= 0:
            page_number = int(self.text_blocks_table.item(current_row, 0).text()) - 1
            if not self.text_blocks_table.item(current_row, 1).text().endswith("K"):
                block_index = int(self.text_blocks_table.item(current_row, 1).text())
            else:
                block_index = self.text_blocks_table.item(current_row, 1).text()

            # print(self.blocks_data)

            if page_number >= 0 and page_number < len(self.blocks_data):
                block = next(
                    (obj for page in self.blocks_data for obj in page
                     if obj["index"] == block_index or (
                             obj.get("text_index") is not None and obj["text_index"] == block_index)),
                    None
                )

                if block in self.blocks_data[page_number]:
                    self.blocks_data[page_number].remove(block)

                for i, block in enumerate(self.blocks_data[page_number], start=self.blocks_data[page_number][0]["index"]):
                    if block["index"] is not None:
                        block["index"] = i

                self.measurement_number = [0]
                for page_blocks in self.blocks_data:
                    for block in page_blocks:
                        if block["index"] is not None:
                            block["index"] = self.measurement_number_calc()

                self.update_blocks_table()
                if page_number == self.current_page:
                    self.draw_highlights()

    def clear_all(self):
        for i in range(len(self.blocks_data)):
            self.blocks_data[i].clear()

            for rect_item, circle_item, line_item, text_item in list(self.highlights[i]):
                try:
                    if rect_item.scene():
                        self.graphics_scene.removeItem(rect_item)
                    if circle_item.scene():
                        self.graphics_scene.removeItem(circle_item)
                    if line_item.scene():
                        self.graphics_scene.removeItem(line_item)
                    if text_item.scene():
                        self.graphics_scene.removeItem(text_item)
                except RuntimeError:
                    pass

            self.highlights[i].clear()

            self.update_blocks_table()

        self.measurement_number = [0]

    def print_documents_block(self):
        excel_default_path = "T:\\Yhteiset\\LAATU\\Mittapöytäkirjapohjat\\MPK_POHJA_PYSTY_2022.xlsx"
        dialog = PrintDialog(default_path=excel_default_path)
        if dialog.exec_() == QDialog.Accepted:
            excel_path = dialog.get_excel_path()
            print(f"Выбранный путь: {excel_path}")
        else:
            print("Выбор отменен")

        self.progress_window = ProgressWindow(self)
        self.progress_window.show()

        try:
            # PDF Save Dialog
            self.save_path, _ = QFileDialog.getSaveFileName(
                self, "Save Highlighted PDF", self.pdf_document_name + "M", "PDF Files (*.pdf)"
            )
            if not self.save_path:
                self.progress_window.close()
                return

            # Extract directory path
            save_dir = os.path.dirname(self.save_path)

            # Path for Excel
            excel_save_path = os.path.join(save_dir, self.pdf_document_name + "M.xlsx")

            self.progress_window.update_progress(20, "Saving PDF...")

            self.print_pdf()

            self.progress_window.update_progress(40, "Saving Excel...")
            self.print_excel(excel_save_path, excel_path)

            self.progress_window.update_progress(100, "Finished")
            self.progress_window.set_compledet()
        except Exception as e:
            self.progress_window.label.setText(f"Error: {e}")
            self.progress_window.ok_button.setEnabled(True)


    def print_pdf(self):

        # pdf
        if not self.pdf_document:
            print("PDF not loaded.")
            return



        pdf_copy = fitz.open()
        for page_index, page_blocks in enumerate(self.blocks_data):
            page = self.pdf_document.load_page(page_index)

            # Create a new page in a copy of the PDF
            new_page = pdf_copy.new_page(width=page.rect.width, height=page.rect.height)
            pixmap = page.get_pixmap()
            new_page.insert_image(page.rect, pixmap=pixmap)

            for block in page_blocks:
                if block["index"] is not None:
                    # draw rect
                    rect = block["rect"]
                    new_page.draw_rect(
                        fitz.Rect(rect[0], rect[1], rect[2], rect[3]),
                        color=(0, 0, 1),  # blue
                        width=1,
                    )

                    if block["circle_position"]:
                        circle_x, circle_y = block["circle_position"]
                    else:
                        circle_x, circle_y = rect[2] + 20, rect[1] - 10

                    # draw line
                    new_page.draw_line(
                        p1=(rect[2], rect[1]),
                        p2=(self.shorten_line_to_circle(rect[2], rect[1], circle_x, circle_y, self.radius)),
                        color=(0, 0, 1),  # blue
                        width=1,
                    )

                    # draw circle
                    new_page.draw_circle(
                        center=(circle_x, circle_y),
                        radius=self.radius,
                        color=(1, 0, 0),  # red
                        width=1
                    )

                    # draw number inside the circle
                    new_page.insert_text(
                        (circle_x - 5, circle_y + 4),
                        str(block["index"]),
                        fontsize=13,
                        color=(1, 0, 0),  # red
                        fontname="helv"
                    )


        # Save new PDF
        try:
            pdf_copy.save(self.save_path)
            pdf_copy.close()
            print(f"PDF saved successfully to {self.save_path}")
        except Exception as e:
            print(f"Error saving PDF: {e}")


    def find_row_with_text(self, workbook, column_index, sheet_name, search_text):

        sheet = workbook[sheet_name]

        # Iterate through the rows in the specified column
        for row in sheet.iter_rows(min_col=column_index, max_col=column_index, values_only=True):
            if row[0] == search_text:
                # If we find the text, we return the line number
                for cell in sheet.iter_rows(min_col=column_index, max_col=column_index):
                    if cell[0].value == search_text:
                        return cell[0].row  # Return the line number

        return None


    def print_excel(self, excel_save_path, excel_malli):


        #path_to_copy = os.path.join(os.getcwd(), "Mittapöytäkirja malli")
        #file_to_copy = os.path.join(path_to_copy, "MPK_POHJA_PYSTY_2022.xlsx")
        file_to_copy = excel_malli

        if not os.path.exists(file_to_copy):
            print(f"Sample file not found: {file_to_copy}")
            return


        shutil.copyfile(file_to_copy, excel_save_path)

        try:
            workbook_find_height = load_workbook(excel_save_path)
            column_index = openpyxl.utils.column_index_from_string("C")
            find_table_height = self.find_row_with_text(workbook_find_height, column_index, "Taul1", "Allekirjoitus")
            mitta_solu = self.find_row_with_text(workbook_find_height, column_index, "Taul1", "Mitta")

        except Exception as e:
            print(f"Error loading sample Excel file (find): {e}")
            return

        table_height = find_table_height - mitta_solu - 1

        count_of_table = math.ceil(len(self.measurement_number)/table_height)

        source_range = f"B1:H{find_table_height}"

        #make copies of the table
        if len(self.measurement_number)>table_height:
            for i in range(count_of_table-1):
                target_start_cell = f"B{(find_table_height*(i+1))+1}"
                self.copy_data_within_excel(excel_save_path, source_range, target_start_cell, find_table_height)


        try:
            # Initialize Excel
            excel = win32.Dispatch("Excel.Application")
            excel.Visible = False
        except Exception as e:
            print(f"Error loading sample Excel file: {e}")
            return

        workbook = excel.Workbooks.Open(excel_save_path)
        sheet = workbook.ActiveSheet

        count = 0
        count2 = 0

        # insert data
        try:
            for page in self.blocks_data:
                for block in page:
                    index = block["index"]
                    text = block["text"]

                    if count2 != 0:
                        if count2 % table_height == 0:
                            count += mitta_solu + 1

                    sheet.Cells(mitta_solu + 1 + count, 2).Value = index
                    sheet.Cells(mitta_solu + 1 + count, 2).NumberFormat = "@" # value like a text
                    sheet.Cells(mitta_solu + 1 + count, 3).Value = text
                    sheet.Cells(mitta_solu + 1 + count, 3).NumberFormat = "@" # value like a text

                    #sheet.cell(row=10 + count, column=1, value=index)
                    #sheet.cell(row=10 + count, column=2, value=text)

                    count += 1
                    count2 += 1
                print("page added")

            print("all data added")
        except Exception as e:
            print(f"Error insert data: {e}")

        # Reset indexes to end
        result = (count2 + 1) % table_height
        if result != 0:
            num = table_height - result
            for i in range(num+1):
                sheet.Cells(mitta_solu + 1 + count, 2).Value = ""
                count += 1

        try:
            workbook.Save()
            workbook.Close()
            excel.Quit()
            print("Successfully")
        except Exception as e:
            workbook.Close()
            excel.Quit()
            print(f"Error saving Excel file: {e}")



    def copy_data_within_excel(self, file_path, source_range, target_start_cell, table_height):
        try:
            # Initialize Excel
            excel = win32.Dispatch("Excel.Application")
            excel.Visible = False

            # Open the file
            workbook = excel.Workbooks.Open(file_path)
            sheet = workbook.ActiveSheet

            for row in sheet.Rows:
                row_number = row.Row
                if row_number > table_height:
                    break
                height = sheet.Rows(row_number).RowHeight

                sheet.Rows(row_number + table_height).RowHeight = height

            # Copy the range
            source_range_object = sheet.Range(source_range)
            source_range_object.Copy()

            # Insert into the target range
            target_cell = sheet.Range(target_start_cell)
            target_cell.Select()
            sheet.Paste()

            # Save and close the file
            workbook.Save()
            workbook.Close()
            excel.Quit()


        except Exception as e:
            print(f"Error copying data: {e}")

    def instruction_block(self):
        self.instruction_window = Instructions()
        self.instruction_window.show()

    def update_block_text(self, row, column):

        if column == 2:
            try:
                page_number = int(self.text_blocks_table.item(row, 0).text()) - 1
                if not self.text_blocks_table.item(row, 1).text().endswith("K"):
                    block_index = int(self.text_blocks_table.item(row, 1).text())
                else:
                    block_index = self.text_blocks_table.item(row, 1).text()
                new_text = self.text_blocks_table.item(row, 2).text()

                block = next(
                    (obj for page in self.blocks_data for obj in page
                     if obj["index"] == block_index or (
                                 obj.get("text_index") is not None and obj["text_index"] == block_index)),
                    None
                )
                if block and block["page"] == page_number:
                    block["text"] = new_text

            except Exception as e:
                print(f"Error updating block text: {e}")

    def add_row_to_table(self):
        # Получить выбранную строку
        current_row = self.text_blocks_table.currentRow()

        # Если ничего не выбрано, добавляем строку в конец
        if current_row == -1:
            current_row = self.text_blocks_table.rowCount() - 1

        # Вставляем новую строку после выбранной
        new_row_position = current_row + 1
        self.text_blocks_table.insertRow(new_row_position)

        # Заполнить столбцы страницы, индекса и текста
        page_item = QTableWidgetItem(str(self.current_page + 1))  # Или задайте значение по умолчанию
        page_item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
        self.text_blocks_table.setItem(new_row_position, 0, page_item)

        index_item = QTableWidgetItem(str(self.measurement_text_calc()) + "K")
        index_item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
        self.text_blocks_table.setItem(new_row_position, 1, index_item)

        text_item = QTableWidgetItem("")
        self.text_blocks_table.setItem(new_row_position, 2, text_item)

        pituus = 0
        for i in range(len(self.blocks_data)):
            pituus += len(self.blocks_data[i])

        # Обновить данные blocks_data
        self.blocks_data[self.current_page].insert(new_row_position - pituus, {
            "page": self.current_page,
            "rect": None,
            "text": "",
            "index": None,
            "circle_position": None,
            "text_index": index_item.text()
        })

    def measurement_text_calc(self):
        self.measurement_text.append(self.measurement_text[-1]+1)
        return self.measurement_text[-1]


if __name__ == "__main__":
    app = QApplication(sys.argv)
    viewer = PDFViewer()
    viewer.show()
    sys.exit(app.exec_())
